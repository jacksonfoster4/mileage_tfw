from preferences import preferences

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

from django.contrib.staticfiles import finders 
from django.conf import settings
from django.apps import apps

from datetime import datetime

import os
from io import BytesIO

import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

class Spreadsheet():
    def __init__(self, user, entries):
        self.email = preferences.CoreAppSettings.spreadsheet_email #pylint: disable=no-member
        self.user = user
        self.entries = entries
        self.offset = 8 # where the entries begin

        self.cell = lambda column, row: "{}{}".format(self.columns[column],row)
        
        self.cells = {
            'reimbursement_rate': 'E3',
            'total_mileage': 'E4',
            'total_reimbursement': 'E5',
            'employee_name': 'B3'
        }

        self.columns = {
            'date': 'A',
            'destination': 'B',
            'notes': 'C',
            'odometer_start': 'D',
            'odometer_end': 'E',
            'mileage': 'F',
            'reimbursement': 'G'
        }

    
        self.wb = Workbook()



    def as_dict(self):
        ws = self.wb.active
        total_miles, total_reimbursement = 0, 0

        final = {'entries': []}

        for entry in self.entries:
            entry_dict = {}
            total_miles += entry.miles_driven()
            total_reimbursement += entry.amount_reimbursed()

            entry_dict['date'] = entry.entry_date.strftime("%m-%d-%Y")
            entry_dict['destination'] = entry.destination
            entry_dict['notes'] = entry.notes
            entry_dict['odometer_start'] = entry.odo_start
            entry_dict['odometer_end'] = entry.odo_end
            entry_dict['mileage'] = entry.miles_driven()
            entry_dict['reimbursement'] = entry.amount_reimbursed()

            final['entries'].append(entry_dict)
        

        final['total_mileage'] = total_miles
        final['total_reimbursement'] = total_reimbursement
        final['reimbursement_rate'] = "${}".format(preferences.CoreAppSettings.reimbursement_rate) # pylint: disable=no-member
        final['employee_name'] = "{} {}".format(self.user.first_name, self.user.last_name)

        return final



    def write_to_spreadsheet(self):
        ws = self.wb.active
        total_miles, total_reimbursement = 0, 0

        for i, entry in enumerate(self.entries):
            row = i + self.offset

            total_miles += entry.miles_driven()
            total_reimbursement += entry.amount_reimbursed()

            ws[self.cell('date',row)] = entry.entry_date.strftime("%m-%d-%Y")
            ws[self.cell('destination',row)] = entry.destination
            ws[self.cell('notes',row)] = entry.notes
            ws[self.cell('odometer_start',row)] = entry.odo_start
            ws[self.cell('odometer_end',row)] = entry.odo_end
            ws[self.cell('mileage',row)] = entry.miles_driven()
            ws[self.cell('reimbursement',row)] = "${}".format(entry.amount_reimbursed())

        ws[self.cells['total_mileage']] = total_miles
        ws[self.cells['total_reimbursement']] = "${}".format(total_reimbursement)
        ws[self.cells['reimbursement_rate']] = "${}".format(preferences.CoreAppSettings.reimbursement_rate) # pylint: disable=no-member
        ws[self.cells['employee_name']] = "{} {}".format(self.user.first_name, self.user.last_name)

        end_of_entries = len(self.entries) + self.offset
        ws["A{}".format(end_of_entries)] = 'Total'
        ws[self.cell('mileage', end_of_entries)] = total_miles
        ws[self.cell('reimbursement', end_of_entries)] = "${}".format(total_reimbursement)


        self.wb.active = ws
        self.add_styles_to_spreadsheet()
        return self.wb.active

    def add_styles_to_spreadsheet(self):

        cells = {
            'A1': 'Mileage Log and Reimbursement',
            'A3': 'Employee Name: ',
            'D3': 'Rate Per Mile',
            'D4': 'Total Mileage',
            'D5': 'Total',
            'A7': 'Date',
            'B7': 'Destination',
            'C7': 'Description/Notes',
            'D7': 'Odometer Start',
            'E7': 'Odometer End',
            'F7': 'Mileage',
            'G7': 'Reimbursement',
        }
        for k, v in cells:
            self.wb.active[k] = v

            
        #reset spreadsheet. colors get messed up for some reason. might be from BytesIO -> .xlsx
        for i, row in enumerate(self.wb.active.rows,1):
            for cell in self.wb.active[i]:
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(bgColor="ffffff", fill_type = "solid")
                cell.font = Font(name='Arial', color="000000", size=14)
                cell.border = Border(
                                left=Side(border_style='thin', color='ffffff'),
                                right=Side(border_style='thin', color='ffffff'),
                                top=Side(border_style='thin', color='ffffff'),
                                bottom=Side(border_style='thin', color='ffffff')
                            ) 

        for i, row in enumerate(self.wb.active.rows,7): # format entries
            if i == len(self.entries)+1: break
            for cell in self.wb.active[i]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Arial', color="000000", size=14)
                cell.border = Border(
                                left=Side(border_style='thin', color='f4f4f4'),
                                right=Side(border_style='thin', color='f4f4f4'),
                                top=Side(border_style='thin', color='f4f4f4'),
                                bottom=Side(border_style='thin', color='f4f4f4')
                            ) 

        title = ['A1', 'B1','C1']
        employee_name = 'B3'
        employee_name_label = 'A3'
        totals_labels = ['D3', 'D4', 'D5']
        totals = ['E3', 'E4', 'E5']
        
        for cell in list(self.wb.active.rows)[6]: # format column titles
            cell.font = Font(name='Arial', bold=True, size=18)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000')
                            ) 
            cell.fill = PatternFill(fgColor="A8A8A8", fill_type = "solid")

        for el in title: # format large spreadsheet title
            self.wb.active[el].font = Font(name='Arial', color='000000', bold=True, size=18)
            self.wb.active[el].alignment = Alignment(horizontal='left')

        # format employee name
        self.wb.active[employee_name_label].font = Font(name='Arial', color="000000", size=14, bold=True)
        self.wb.active[employee_name].border = Border(bottom=Side(border_style='thin', color='000000'))   

        for el in totals_labels: # format totals
            self.wb.active[el].font = Font(bold=True, size=12)
        for el in totals:
            self.wb.active[el].border = Border(
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000')
                            ) 
        
        for cell in self.wb.active[len(self.entries)+self.offset]: # format total row at the end of entries
            cell.font = Font(name='Arial', color='000000', bold=True, size=18)
        
                        
                                      
    def save_as_binary_stream(self):
        stream = BytesIO()
        self.wb.save(stream)
        return stream
    


    def email_spreadsheet(self):
        port = 465  # For SSL

        sender_email = os.getenv("SENDER_EMAIL")
        password = os.getenv("SENDER_EMAIL_PASSWORD")
        receiver_email = preferences.CoreAppSettings.spreadsheet_email # pylint: disable=no-member

        message = MIMEMultipart("alternative")
        message["Subject"] = "multipart test"
        message["From"] = sender_email
        message["To"] = receiver_email

        text = """\
                Hi,
                How are you?
                Real Python has many great tutorials:
                www.realpython.com"""
        html = """\
                <html>
                <body>
                    <p>Hi,<br>
                    How are you?<br>
                    <a style='color:red'href="http://www.realpython.com">Real Python</a> 
                    has many great tutorials.
                    </p>
                </body>
                </html>
                """
        part1 = MIMEText(text, "plain")
        part2 = MIMEText(html, "html")
        
        part = MIMEBase("application", "octet-stream")
        part.set_payload(self.save_as_binary_stream().getvalue())

        # Encode file in ASCII characters to send by email    
        encoders.encode_base64(part)

        # Add header as key/value pair to attachment part
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {self.wb.active.title}.xlsx",
        )

        # Add attachment to message and convert message to string
        message.attach(part)
        text = message.as_string()

        message.attach(part1)
        message.attach(part2)
        # Create a secure SSL context
        context = ssl.create_default_context()

        with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, message.as_string())




